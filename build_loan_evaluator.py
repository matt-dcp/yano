#!/usr/bin/env python3
"""Build Casa Yano loan evaluator spreadsheet for evaluating lender options
and identifying valuable concessions."""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

# Load latest pro forma data
with open(Path(__file__).parent / "public" / "data.js") as f:
    raw = f.read()
    D = json.loads(raw.split("window.__DATA__ = ")[1].rstrip(";\n"))

PF = D["proForma"]
CURRENT_NOI = PF["noiAfterKnown"]

# Styles
DARK = PatternFill("solid", fgColor="2D2D2D")
GOLD_FILL = PatternFill("solid", fgColor="C5A55A")
LIGHT_GOLD = PatternFill("solid", fgColor="F5ECD7")
LIGHT_GRAY = PatternFill("solid", fgColor="F5F5F5")
INPUT_FILL = PatternFill("solid", fgColor="DCE6F1")
GREEN_FILL = PatternFill("solid", fgColor="E8F5E9")
YELLOW_FILL = PatternFill("solid", fgColor="FFFDE7")
RED_FILL = PatternFill("solid", fgColor="FFEBEE")

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
GOLD_FONT = Font(name="Arial", bold=True, color="C5A55A", size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=16, color="2D2D2D")
SECTION_FONT = Font(name="Arial", bold=True, size=12, color="2D2D2D")
BODY = Font(name="Arial", size=10, color="333333")
BOLD = Font(name="Arial", bold=True, size=10, color="333333")
INPUT_FONT = Font(name="Arial", size=10, color="0000FF", bold=True)
NOTE_FONT = Font(name="Arial", size=9, color="999999", italic=True)

MONEY = '$#,##0;($#,##0);"-"'
MONEY_K = '$#,##0,"K"'
PCT = '0.00%'
PCT0 = '0.0%'
NUM = '#,##0'
NUM2 = '0.00'
DSCR_FMT = '0.00"x"'

THIN_B = Border(bottom=Side(style="thin", color="DDDDDD"))
THICK_B = Border(bottom=Side(style="medium", color="2D2D2D"))
ALL_THIN = Border(left=Side(style="thin", color="DDDDDD"),
                   right=Side(style="thin", color="DDDDDD"),
                   top=Side(style="thin", color="DDDDDD"),
                   bottom=Side(style="thin", color="DDDDDD"))


def hdr_row(ws, row, start_col, end_col):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = DARK
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def section_header(ws, row, title, end_col=8):
    ws.cell(row, 1, title).font = SECTION_FONT
    for c in range(1, end_col + 1):
        ws.cell(row, c).border = THICK_B
    return row + 1


wb = Workbook()


# ════════════════════════════════════════════════════════
# TAB 1: Dashboard / Assumptions
# ════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Dashboard"
ws1.sheet_properties.tabColor = "C5A55A"

for col, w in [("A", 32), ("B", 18), ("C", 4), ("D", 32), ("E", 18)]:
    ws1.column_dimensions[col].width = w

r = 1
ws1.cell(r, 1, "Casa Yano - Loan Evaluator").font = Font(name="Arial", bold=True, size=18, color="2D2D2D")
ws1.merge_cells("A1:E1")
r = 2
ws1.cell(r, 1, "Identify optimal loan structure and quantify value of lender concessions").font = NOTE_FONT
ws1.merge_cells("A2:E2")
r = 3
ws1.cell(r, 1, f"Built from {D['generatedAt'][:10]} pro forma. Blue cells are inputs - adjust to test scenarios.").font = NOTE_FONT
ws1.merge_cells("A3:E3")

# ── INPUTS section ──
r = 5
r = section_header(ws1, r, "PRO FORMA INPUTS", end_col=2)

# Cell B6 = NOI (master input)
ws1.cell(r, 1, "Projected 2026 NOI (after fixed costs)").font = BODY
ws1.cell(r, 2, CURRENT_NOI).font = INPUT_FONT
ws1.cell(r, 2).fill = INPUT_FILL
ws1.cell(r, 2).number_format = MONEY
NOI_REF = f"Dashboard!$B${r}"
NOI_ROW = r
r += 1

# NOI scenarios
ws1.cell(r, 1, "  NOI scenario adjustment").font = BODY
ws1.cell(r, 2, 1.00).font = INPUT_FONT
ws1.cell(r, 2).fill = INPUT_FILL
ws1.cell(r, 2).number_format = '0.0%'
NOI_ADJ_REF = f"Dashboard!$B${r}"
ws1.cell(r, 5, "(0% = current pro forma; +10% = optimistic case)").font = NOTE_FONT
r += 1

ws1.cell(r, 1, "  Adjusted NOI").font = BOLD
ws1.cell(r, 2, f"=B{NOI_ROW}*B{NOI_ROW+1}").font = BOLD
ws1.cell(r, 2).number_format = MONEY
ws1.cell(r, 2).fill = LIGHT_GOLD
ADJ_NOI_REF = f"Dashboard!$B${r}"
ADJ_NOI_ROW = r
r += 2

# ── Property assumptions ──
r = section_header(ws1, r, "PROPERTY ASSUMPTIONS", end_col=2)
ws1.cell(r, 1, "Total Basis (per QBO)").font = BODY
ws1.cell(r, 2, 4_200_000).font = INPUT_FONT
ws1.cell(r, 2).fill = INPUT_FILL
ws1.cell(r, 2).number_format = MONEY
BASIS_REF = f"Dashboard!$B${r}"
r += 1
ws1.cell(r, 1, "Existing Debt (private loan, 9.5% IO)").font = BODY
ws1.cell(r, 2, 1_500_000).font = INPUT_FONT
ws1.cell(r, 2).fill = INPUT_FILL
ws1.cell(r, 2).number_format = MONEY
EXISTING_DEBT_REF = f"Dashboard!$B${r}"
r += 1
ws1.cell(r, 1, "Existing Equity").font = BODY
ws1.cell(r, 2, f"={BASIS_REF}-{EXISTING_DEBT_REF}").font = BOLD
ws1.cell(r, 2).number_format = MONEY
r += 2

# ── Cap rate / LTV / DSCR globals ──
r = section_header(ws1, r, "VALUATION & UNDERWRITING ASSUMPTIONS", end_col=2)
ws1.cell(r, 1, "Cap Rate (for valuation - center of sensitivity)").font = BODY
ws1.cell(r, 2, 0.06).font = INPUT_FONT
ws1.cell(r, 2).fill = INPUT_FILL
ws1.cell(r, 2).number_format = PCT
CAP_RATE_REF = f"Dashboard!$B${r}"
r += 1
ws1.cell(r, 1, "Implied Property Value").font = BOLD
ws1.cell(r, 2, f"={ADJ_NOI_REF}/{CAP_RATE_REF}").font = BOLD
ws1.cell(r, 2).number_format = MONEY
ws1.cell(r, 2).fill = LIGHT_GOLD
VALUE_REF = f"Dashboard!$B${r}"
r += 2

# ── Lender Scenarios summary ──
r = section_header(ws1, r, "LENDER SCENARIOS (CURRENT TERMS)", end_col=5)
hdr_row(ws1, r, 1, 5)
ws1.cell(r, 1, "Lender / Option")
ws1.cell(r, 2, "Loan Amount")
ws1.cell(r, 4, "Rate")
ws1.cell(r, 5, "Amort")
r += 1

scenarios = [
    ("Ben / MBT - Option 1", 2_300_000, "Same rate", 0.0625, 25, "1.35x DSCR / 55% LTV / 0.50% fee / Day-1 amort"),
    ("Ben / MBT - Option 2 (with earn-out)", 2_800_000, "with earn-out", 0.0640, 25, "1.35x DSCR / 60% LTV / 0.75% fee / 12mo IO"),
    ("ServisFirst (current indicative)", 0, "TBD", 0.0650, 20, "1.50x DSCR / 75% LTV / Earn-out 2-phase TBD"),
]
ws1.cell(r-1, 3, "Notes")
ws1.column_dimensions["C"].width = 35
for label, loan, note_extra, rate, amort, notes in scenarios:
    ws1.cell(r, 1, label).font = BODY
    if loan > 0:
        ws1.cell(r, 2, loan).number_format = MONEY
    ws1.cell(r, 2).font = BODY
    ws1.cell(r, 3, notes).font = NOTE_FONT
    ws1.cell(r, 4, rate).number_format = PCT
    ws1.cell(r, 4).font = BODY
    ws1.cell(r, 5, amort).font = BODY
    for c in range(1, 6):
        ws1.cell(r, c).border = THIN_B
    r += 1

r += 2

# ── Quick-glance KPIs ──
r = section_header(ws1, r, "DECISION SUPPORT - SEE OTHER TABS FOR DETAIL", end_col=5)
ws1.cell(r, 1, "Tab 2: Valuation Sensitivity").font = BOLD
ws1.cell(r, 2, "Property value at various NOI / cap rate combinations").font = NOTE_FONT
ws1.merge_cells(f"B{r}:E{r}")
r += 1
ws1.cell(r, 1, "Tab 3: Max Loan Capacity").font = BOLD
ws1.cell(r, 2, "Maximum loan by LTV constraint, DSCR constraint, and binding factor").font = NOTE_FONT
ws1.merge_cells(f"B{r}:E{r}")
r += 1
ws1.cell(r, 1, "Tab 4: Lender Comparison").font = BOLD
ws1.cell(r, 2, "Side-by-side analysis of all three lender scenarios").font = NOTE_FONT
ws1.merge_cells(f"B{r}:E{r}")
r += 1
ws1.cell(r, 1, "Tab 5: Concession Levers").font = BOLD
ws1.cell(r, 2, "How much each lender concession is worth in additional proceeds").font = NOTE_FONT
ws1.merge_cells(f"B{r}:E{r}")

ws1.freeze_panes = "A5"


# ════════════════════════════════════════════════════════
# TAB 2: Valuation Sensitivity
# ════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Valuation Sensitivity")
ws2.sheet_properties.tabColor = "2D8B2D"

ws2.column_dimensions["A"].width = 26
for c in range(2, 9):
    ws2.column_dimensions[get_column_letter(c)].width = 16

r = 1
ws2.cell(r, 1, "Property Valuation Sensitivity").font = TITLE_FONT
ws2.merge_cells("A1:H1")
r = 2
ws2.cell(r, 1, "Implied property value at various NOI scenarios and cap rates").font = NOTE_FONT
ws2.merge_cells("A2:H2")
r = 3
ws2.cell(r, 1, "Formula: Property Value = NOI / Cap Rate").font = NOTE_FONT
ws2.merge_cells("A3:H3")

r = 5
# Row of cap rates as headers
cap_rates = [0.045, 0.050, 0.055, 0.060, 0.065, 0.070, 0.075]
ws2.cell(r, 1, "NOI Scenario").font = HEADER_FONT
ws2.cell(r, 1).fill = DARK
ws2.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")
for i, cr in enumerate(cap_rates):
    cell = ws2.cell(r, i+2, cr)
    cell.fill = DARK
    cell.font = HEADER_FONT
    cell.number_format = PCT
    cell.alignment = Alignment(horizontal="center")
r += 1

# NOI scenarios
noi_scenarios = [
    ("Conservative (-10%)", 0.90),
    ("Conservative (-5%)", 0.95),
    ("Current Pro Forma", 1.00),
    ("Modest Beat (+5%)", 1.05),
    ("Strong Beat (+10%)", 1.10),
    ("Bull Case (+15%)", 1.15),
    ("Bull Case (+20%)", 1.20),
]
val_start_row = r
for label, mult in noi_scenarios:
    ws2.cell(r, 1, label).font = BOLD
    if mult == 1.00:
        ws2.cell(r, 1).fill = LIGHT_GOLD
    for i, cr in enumerate(cap_rates):
        cell = ws2.cell(r, i+2)
        cell.value = f"={NOI_REF}*{mult}/{cr}"
        cell.number_format = MONEY
        cell.font = BODY
        if mult == 1.00:
            cell.fill = LIGHT_GOLD
        cell.border = THIN_B
    r += 1
val_end_row = r - 1

# Color scale
last_col = get_column_letter(len(cap_rates) + 1)
rng = f"B{val_start_row}:{last_col}{val_end_row}"
ws2.conditional_formatting.add(rng, ColorScaleRule(
    start_type='min', start_color='FFEBEE',
    mid_type='percentile', mid_value=50, mid_color='FFFDE7',
    end_type='max', end_color='E8F5E9'))

r += 2
ws2.cell(r, 1, "NOTES").font = SECTION_FONT
r += 1
ws2.cell(r, 1, "- 6.0% cap rate is reasonable for SB STR with strong location and renovation").font = NOTE_FONT
ws2.merge_cells(f"A{r}:H{r}")
r += 1
ws2.cell(r, 1, "- Lender appraisers typically use 5.5% - 6.5% range for SB lodging assets").font = NOTE_FONT
ws2.merge_cells(f"A{r}:H{r}")
r += 1
ws2.cell(r, 1, "- Higher cap rate = more conservative valuation (lower value)").font = NOTE_FONT
ws2.merge_cells(f"A{r}:H{r}")

ws2.freeze_panes = "B6"


# ════════════════════════════════════════════════════════
# TAB 3: Max Loan Capacity
# ════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Max Loan Capacity")
ws3.sheet_properties.tabColor = "1565C0"

ws3.column_dimensions["A"].width = 30
for c in range(2, 11):
    ws3.column_dimensions[get_column_letter(c)].width = 16

r = 1
ws3.cell(r, 1, "Maximum Loan Capacity Analysis").font = TITLE_FONT
ws3.merge_cells("A1:J1")
r = 2
ws3.cell(r, 1, "Max loan by LTV constraint, DSCR constraint, and binding factor (lesser of two)").font = NOTE_FONT
ws3.merge_cells("A2:J2")
r = 4

# ── Section 1: Max loan by LTV ──
ws3.cell(r, 1, "MAX LOAN BY LTV (using current pro forma NOI)").font = SECTION_FONT
ws3.cell(r, 1).border = THICK_B
for c in range(2, 11):
    ws3.cell(r, c).border = THICK_B
r += 1

ws3.cell(r, 1, "Cap Rate Used").fill = DARK
ws3.cell(r, 1).font = HEADER_FONT
ws3.cell(r, 1).alignment = Alignment(horizontal="center")
ltv_ratios = [0.50, 0.55, 0.60, 0.65, 0.70, 0.75]
for i, ltv in enumerate(ltv_ratios):
    cell = ws3.cell(r, i+2, f"{ltv*100:.0f}% LTV")
    cell.fill = DARK
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center")
r += 1

cap_rates_for_loan = [0.055, 0.060, 0.065]
ltv_start_r = r
for cr in cap_rates_for_loan:
    ws3.cell(r, 1, f"{cr*100:.1f}% cap").font = BOLD
    for i, ltv in enumerate(ltv_ratios):
        # Max loan = NOI / cap_rate * LTV
        cell = ws3.cell(r, i+2)
        cell.value = f"={ADJ_NOI_REF}/{cr}*{ltv}"
        cell.number_format = MONEY
        cell.border = THIN_B
        cell.font = BODY
    r += 1

r += 2

# ── Section 2: Max loan by DSCR ──
ws3.cell(r, 1, "MAX LOAN BY DSCR (using current pro forma NOI)").font = SECTION_FONT
ws3.cell(r, 1).border = THICK_B
for c in range(2, 11):
    ws3.cell(r, c).border = THICK_B
r += 1

dscr_mins = [1.20, 1.25, 1.35, 1.50]
ws3.cell(r, 1, "Rate / Amort").fill = DARK
ws3.cell(r, 1).font = HEADER_FONT
ws3.cell(r, 1).alignment = Alignment(horizontal="center")
for i, d in enumerate(dscr_mins):
    cell = ws3.cell(r, i+2, f"{d:.2f}x DSCR")
    cell.fill = DARK
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center")
r += 1

# Rate / amort combinations
rate_amort_combos = [
    (0.0625, 25, "6.25% / 25-yr"),
    (0.0640, 25, "6.40% / 25-yr"),
    (0.0625, 30, "6.25% / 30-yr"),
    (0.0650, 20, "6.50% / 20-yr"),
    (0.0650, 25, "6.50% / 25-yr"),
    (0.0650, 30, "6.50% / 30-yr"),
    (0.0700, 25, "7.00% / 25-yr"),
    (0.0750, 25, "7.50% / 25-yr"),
]

# Max loan formula: NOI / DSCR / annual_constant
# Where annual_constant = monthly_pmt_factor * 12 = [r(1+r)^n / ((1+r)^n - 1)] * 12
def annual_constant_formula(rate, years):
    """Excel formula for annual debt constant given rate and years"""
    r = rate / 12
    n = years * 12
    monthly = r * (1+r)**n / ((1+r)**n - 1)
    return monthly * 12

dscr_start_r = r
for rate, amort, label in rate_amort_combos:
    ws3.cell(r, 1, label).font = BOLD
    ac = annual_constant_formula(rate, amort)
    for i, dscr in enumerate(dscr_mins):
        cell = ws3.cell(r, i+2)
        cell.value = f"={ADJ_NOI_REF}/{dscr}/{ac}"
        cell.number_format = MONEY
        cell.border = THIN_B
        cell.font = BODY
    r += 1

r += 2

# ── Section 3: Binding constraint ──
ws3.cell(r, 1, "BINDING CONSTRAINT (lesser of LTV or DSCR max)").font = SECTION_FONT
ws3.cell(r, 1).border = THICK_B
for c in range(2, 11):
    ws3.cell(r, c).border = THICK_B
r += 1

ws3.cell(r, 1, "Selected lender scenarios applied to current pro forma:").font = NOTE_FONT
ws3.merge_cells(f"A{r}:E{r}")
r += 1

ws3.cell(r, 1, "Lender Scenario").fill = DARK
ws3.cell(r, 1).font = HEADER_FONT
ws3.cell(r, 2, "LTV Max").fill = DARK
ws3.cell(r, 2).font = HEADER_FONT
ws3.cell(r, 2).alignment = Alignment(horizontal="center")
ws3.cell(r, 3, "DSCR Max").fill = DARK
ws3.cell(r, 3).font = HEADER_FONT
ws3.cell(r, 3).alignment = Alignment(horizontal="center")
ws3.cell(r, 4, "Binding").fill = DARK
ws3.cell(r, 4).font = HEADER_FONT
ws3.cell(r, 4).alignment = Alignment(horizontal="center")
ws3.cell(r, 5, "Constraint").fill = DARK
ws3.cell(r, 5).font = HEADER_FONT
ws3.cell(r, 5).alignment = Alignment(horizontal="center")
r += 1

# Each lender's binding constraints
lender_constraints = [
    ("Ben Option 1: 55% LTV / 1.35x / 6.25% / 25-yr", 0.55, 1.35, 0.0625, 25),
    ("Ben Option 2: 60% LTV / 1.35x / 6.40% / 25-yr", 0.60, 1.35, 0.0640, 25),
    ("ServisFirst: 75% LTV / 1.50x / 6.50% / 20-yr", 0.75, 1.50, 0.0650, 20),
    ("ServisFirst neg: 75% LTV / 1.35x / 6.50% / 25-yr", 0.75, 1.35, 0.0650, 25),
]
for label, ltv, dscr, rate, amort in lender_constraints:
    ac = annual_constant_formula(rate, amort)
    cap = 0.06  # Use 6% cap for value calculation

    ws3.cell(r, 1, label).font = BODY
    ws3.cell(r, 2).value = f"={ADJ_NOI_REF}/{cap}*{ltv}"
    ws3.cell(r, 2).number_format = MONEY
    ws3.cell(r, 3).value = f"={ADJ_NOI_REF}/{dscr}/{ac}"
    ws3.cell(r, 3).number_format = MONEY
    ws3.cell(r, 4).value = f"=MIN(B{r},C{r})"
    ws3.cell(r, 4).number_format = MONEY
    ws3.cell(r, 4).font = BOLD
    ws3.cell(r, 4).fill = LIGHT_GOLD
    ws3.cell(r, 5).value = f'=IF(B{r}<C{r},"LTV","DSCR")'
    ws3.cell(r, 5).font = BODY
    ws3.cell(r, 5).alignment = Alignment(horizontal="center")
    for c in range(1, 6):
        ws3.cell(r, c).border = THIN_B
    r += 1

r += 1
ws3.cell(r, 1, "Note: LTV calc assumes 6.0% cap rate. Adjust on Dashboard tab to test different cap rates.").font = NOTE_FONT
ws3.merge_cells(f"A{r}:H{r}")

ws3.freeze_panes = "A5"


# ════════════════════════════════════════════════════════
# TAB 4: Lender Comparison
# ════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Lender Comparison")
ws4.sheet_properties.tabColor = "7B1FA2"

ws4.column_dimensions["A"].width = 32
for c in range(2, 6):
    ws4.column_dimensions[get_column_letter(c)].width = 19

r = 1
ws4.cell(r, 1, "Lender Comparison").font = TITLE_FONT
ws4.merge_cells("A1:E1")
r = 2
ws4.cell(r, 1, "Side-by-side at current terms with cash flow and return analysis").font = NOTE_FONT
ws4.merge_cells("A2:E2")
r = 4

# Header row with lenders
ws4.cell(r, 1, "").fill = DARK
hdrs = ["Ben Opt 1", "Ben Opt 2", "ServisFirst (current)", "ServisFirst (negotiated)"]
for i, h in enumerate(hdrs):
    cell = ws4.cell(r, i+2, h)
    cell.fill = DARK
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
r += 1

# Loan terms
ws4.cell(r, 1, "LOAN TERMS").font = BOLD
ws4.cell(r, 1).fill = LIGHT_GRAY
for c in range(2, 6):
    ws4.cell(r, c).fill = LIGHT_GRAY
r += 1

scenarios_compare = [
    {"name": "Ben Opt 1", "loan": 2_300_000, "rate": 0.0625, "amort": 25, "fee": 0.005, "io_months": 0, "ltv_cap": 0.55, "dscr_min": 1.35},
    {"name": "Ben Opt 2", "loan": 2_800_000, "rate": 0.0640, "amort": 25, "fee": 0.0075, "io_months": 12, "ltv_cap": 0.60, "dscr_min": 1.35},
    {"name": "ServisFirst (current)", "loan": 0, "rate": 0.0650, "amort": 20, "fee": 0.01, "io_months": 0, "ltv_cap": 0.75, "dscr_min": 1.50},
    {"name": "ServisFirst (negotiated)", "loan": 0, "rate": 0.0625, "amort": 25, "fee": 0.0075, "io_months": 12, "ltv_cap": 0.75, "dscr_min": 1.35},
]

# For ServisFirst, calc max loan from binding constraint at current NOI assuming 6% cap
def annual_const(rate, amort):
    r = rate / 12; n = amort * 12
    return r * (1+r)**n / ((1+r)**n - 1) * 12

for s in scenarios_compare:
    if s["loan"] == 0:
        # Calc max loan as min of LTV and DSCR
        ac = annual_const(s["rate"], s["amort"])
        ltv_max = CURRENT_NOI / 0.06 * s["ltv_cap"]
        dscr_max = CURRENT_NOI / s["dscr_min"] / ac
        s["loan"] = min(ltv_max, dscr_max)

# Loan amount
ws4.cell(r, 1, "Loan Amount").font = BODY
for i, s in enumerate(scenarios_compare):
    cell = ws4.cell(r, i+2, s["loan"])
    cell.number_format = MONEY
    cell.font = BOLD
    cell.alignment = Alignment(horizontal="right")
r += 1
loan_row = r - 1

# Rate
ws4.cell(r, 1, "Interest Rate").font = BODY
for i, s in enumerate(scenarios_compare):
    cell = ws4.cell(r, i+2, s["rate"])
    cell.number_format = PCT
    cell.alignment = Alignment(horizontal="right")
r += 1

# Amort
ws4.cell(r, 1, "Amortization (years)").font = BODY
for i, s in enumerate(scenarios_compare):
    cell = ws4.cell(r, i+2, s["amort"])
    cell.alignment = Alignment(horizontal="right")
r += 1

# IO period
ws4.cell(r, 1, "IO Period (months)").font = BODY
for i, s in enumerate(scenarios_compare):
    cell = ws4.cell(r, i+2, s["io_months"])
    cell.alignment = Alignment(horizontal="right")
r += 1

# Loan fee
ws4.cell(r, 1, "Loan Fee %").font = BODY
for i, s in enumerate(scenarios_compare):
    cell = ws4.cell(r, i+2, s["fee"])
    cell.number_format = PCT
    cell.alignment = Alignment(horizontal="right")
r += 1

# DSCR floor
ws4.cell(r, 1, "DSCR Minimum").font = BODY
for i, s in enumerate(scenarios_compare):
    cell = ws4.cell(r, i+2, s["dscr_min"])
    cell.number_format = DSCR_FMT
    cell.alignment = Alignment(horizontal="right")
r += 1

# LTV cap
ws4.cell(r, 1, "LTV Maximum").font = BODY
for i, s in enumerate(scenarios_compare):
    cell = ws4.cell(r, i+2, s["ltv_cap"])
    cell.number_format = PCT0
    cell.alignment = Alignment(horizontal="right")
r += 2

# Cash flow analysis
ws4.cell(r, 1, "ANNUAL CASH FLOW (Steady-State)").font = BOLD
ws4.cell(r, 1).fill = LIGHT_GRAY
for c in range(2, 6):
    ws4.cell(r, c).fill = LIGHT_GRAY
r += 1

# Annual debt service (steady state, after IO)
ws4.cell(r, 1, "Annual Debt Service").font = BODY
for i, s in enumerate(scenarios_compare):
    ac = annual_const(s["rate"], s["amort"])
    ds = s["loan"] * ac
    cell = ws4.cell(r, i+2, ds)
    cell.number_format = MONEY
    cell.alignment = Alignment(horizontal="right")
r += 1
ds_row = r - 1

# NOI reference
ws4.cell(r, 1, "Projected NOI").font = BODY
for i in range(4):
    cell = ws4.cell(r, i+2)
    cell.value = f"={ADJ_NOI_REF}"
    cell.number_format = MONEY
    cell.alignment = Alignment(horizontal="right")
r += 1
noi_row = r - 1

# CF after DS
ws4.cell(r, 1, "Cash Flow After Debt Service").font = BOLD
for i in range(4):
    cell = ws4.cell(r, i+2)
    col = get_column_letter(i+2)
    cell.value = f"={col}{noi_row}-{col}{ds_row}"
    cell.number_format = MONEY
    cell.font = BOLD
    cell.fill = LIGHT_GOLD
    cell.alignment = Alignment(horizontal="right")
r += 1

# DSCR achieved
ws4.cell(r, 1, "DSCR Achieved").font = BODY
for i in range(4):
    cell = ws4.cell(r, i+2)
    col = get_column_letter(i+2)
    cell.value = f"={col}{noi_row}/{col}{ds_row}"
    cell.number_format = DSCR_FMT
    cell.alignment = Alignment(horizontal="right")
r += 1

# Debt yield
ws4.cell(r, 1, "Debt Yield (NOI / Loan)").font = BODY
for i in range(4):
    cell = ws4.cell(r, i+2)
    col = get_column_letter(i+2)
    cell.value = f"={col}{noi_row}/{col}{loan_row}"
    cell.number_format = PCT
    cell.alignment = Alignment(horizontal="right")
r += 2

# Cash to owner
ws4.cell(r, 1, "CASH OUT AT CLOSE").font = BOLD
ws4.cell(r, 1).fill = LIGHT_GRAY
for c in range(2, 6):
    ws4.cell(r, c).fill = LIGHT_GRAY
r += 1

ws4.cell(r, 1, "Loan Amount").font = BODY
for i in range(4):
    col = get_column_letter(i+2)
    cell = ws4.cell(r, i+2)
    cell.value = f"={col}{loan_row}"
    cell.number_format = MONEY
    cell.alignment = Alignment(horizontal="right")
loan_ref_row = r
r += 1

ws4.cell(r, 1, "Less: Existing Debt Payoff").font = BODY
for i in range(4):
    cell = ws4.cell(r, i+2)
    cell.value = f"=-{EXISTING_DEBT_REF}"
    cell.number_format = MONEY
    cell.alignment = Alignment(horizontal="right")
r += 1

ws4.cell(r, 1, "Less: Loan Fee").font = BODY
for i, s in enumerate(scenarios_compare):
    col = get_column_letter(i+2)
    cell = ws4.cell(r, i+2)
    cell.value = f"=-{col}{loan_ref_row}*{s['fee']}"
    cell.number_format = MONEY
    cell.alignment = Alignment(horizontal="right")
fee_row = r
r += 1

ws4.cell(r, 1, "Less: Doc + 3rd Party Costs (est)").font = BODY
for i in range(4):
    cell = ws4.cell(r, i+2, -31_000)
    cell.number_format = MONEY
    cell.alignment = Alignment(horizontal="right")
costs_row = r
r += 1

ws4.cell(r, 1, "Net Cash to Owner at Close").font = BOLD
for i in range(4):
    col = get_column_letter(i+2)
    cell = ws4.cell(r, i+2)
    cell.value = f"={col}{loan_ref_row}+{col}{loan_ref_row+1}+{col}{loan_ref_row+2}+{col}{loan_ref_row+3}"
    cell.number_format = MONEY
    cell.font = BOLD
    cell.fill = LIGHT_GOLD
    cell.alignment = Alignment(horizontal="right")
r += 2

# Returns
ws4.cell(r, 1, "RETURNS").font = BOLD
ws4.cell(r, 1).fill = LIGHT_GRAY
for c in range(2, 6):
    ws4.cell(r, c).fill = LIGHT_GRAY
r += 1

ws4.cell(r, 1, "Equity Remaining (Basis - Loan)").font = BODY
for i in range(4):
    col = get_column_letter(i+2)
    cell = ws4.cell(r, i+2)
    cell.value = f"={BASIS_REF}-{col}{loan_ref_row}"
    cell.number_format = MONEY
    cell.alignment = Alignment(horizontal="right")
eq_row = r
r += 1

ws4.cell(r, 1, "Cash-on-Cash Return (Steady-State)").font = BOLD
for i in range(4):
    col = get_column_letter(i+2)
    cell = ws4.cell(r, i+2)
    # CF / Equity remaining
    cell.value = f"=({col}{noi_row}-{col}{ds_row})/{col}{eq_row}"
    cell.number_format = PCT
    cell.font = BOLD
    cell.alignment = Alignment(horizontal="right")
r += 1

ws4.cell(r, 1, "Leverage (Loan / Basis)").font = BODY
for i in range(4):
    col = get_column_letter(i+2)
    cell = ws4.cell(r, i+2)
    cell.value = f"={col}{loan_ref_row}/{BASIS_REF}"
    cell.number_format = PCT0
    cell.alignment = Alignment(horizontal="right")
r += 2

ws4.cell(r, 1, "NOTES:").font = SECTION_FONT
r += 1
notes = [
    "- ServisFirst (current): max loan calculated as binding minimum of 75% LTV (at 6% cap) and 1.50x DSCR / 6.50% / 20-yr",
    "- ServisFirst (negotiated): assumes ServisFirst gives on rate (-25 bps), amort (+5 yrs), and DSCR floor (-15 bps)",
    "- Cash flow analysis is steady-state (post-IO if applicable). Year 1 may differ for Ben Option 2 due to 12-month IO.",
    "- All scenarios assume current pro forma NOI; adjust on Dashboard tab to test sensitivities.",
]
for n in notes:
    ws4.cell(r, 1, n).font = NOTE_FONT
    ws4.merge_cells(f"A{r}:E{r}")
    r += 1

ws4.freeze_panes = "B5"


# ════════════════════════════════════════════════════════
# TAB 5: Concession Levers
# ════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Concession Levers")
ws5.sheet_properties.tabColor = "E91E63"

ws5.column_dimensions["A"].width = 36
for c in range(2, 7):
    ws5.column_dimensions[get_column_letter(c)].width = 16

r = 1
ws5.cell(r, 1, "Concession Levers - What Each Concession is Worth").font = TITLE_FONT
ws5.merge_cells("A1:F1")
r = 2
ws5.cell(r, 1, "Quantifies how much additional loan capacity each lender concession unlocks").font = NOTE_FONT
ws5.merge_cells("A2:F2")
r = 4

# Section: Ben Option 2 - what each concession is worth
ws5.cell(r, 1, "BEN / MBT - OPTION 2 CONCESSION ANALYSIS").font = SECTION_FONT
ws5.cell(r, 1).border = THICK_B
for c in range(2, 7):
    ws5.cell(r, c).border = THICK_B
r += 1

ws5.cell(r, 1, "Base case: $2.8M / 6.40% / 25-yr / 1.35x DSCR / 60% LTV").font = NOTE_FONT
ws5.merge_cells(f"A{r}:F{r}")
r += 2

# Header row
ws5.cell(r, 1, "Concession").fill = DARK
ws5.cell(r, 1).font = HEADER_FONT
ws5.cell(r, 2, "Change").fill = DARK
ws5.cell(r, 2).font = HEADER_FONT
ws5.cell(r, 2).alignment = Alignment(horizontal="center")
ws5.cell(r, 3, "New Max Loan").fill = DARK
ws5.cell(r, 3).font = HEADER_FONT
ws5.cell(r, 3).alignment = Alignment(horizontal="center")
ws5.cell(r, 4, "vs Base $2.8M").fill = DARK
ws5.cell(r, 4).font = HEADER_FONT
ws5.cell(r, 4).alignment = Alignment(horizontal="center")
ws5.cell(r, 5, "$ Value").fill = DARK
ws5.cell(r, 5).font = HEADER_FONT
ws5.cell(r, 5).alignment = Alignment(horizontal="center")
ws5.cell(r, 6, "% Move").fill = DARK
ws5.cell(r, 6).font = HEADER_FONT
ws5.cell(r, 6).alignment = Alignment(horizontal="center")
r += 1

# Base case
ben_base_loan = 2_800_000
ben_base_rate = 0.0640
ben_base_amort = 25
ben_base_dscr = 1.35
ben_base_ltv = 0.60
cap = 0.06

def calc_max_loan(rate, amort, dscr, ltv, noi=CURRENT_NOI, cap_rate=cap):
    ac = annual_const(rate, amort)
    ltv_max = noi / cap_rate * ltv
    dscr_max = noi / dscr / ac
    return min(ltv_max, dscr_max), ltv_max, dscr_max

base_loan, base_ltv_max, base_dscr_max = calc_max_loan(ben_base_rate, ben_base_amort, ben_base_dscr, ben_base_ltv)
ws5.cell(r, 1, "BASE CASE").font = BOLD
ws5.cell(r, 1).fill = LIGHT_GOLD
ws5.cell(r, 2, "-").alignment = Alignment(horizontal="center")
ws5.cell(r, 3, base_loan).number_format = MONEY
ws5.cell(r, 3).font = BOLD
ws5.cell(r, 3).fill = LIGHT_GOLD
ws5.cell(r, 4, 0).number_format = MONEY
ws5.cell(r, 5, "-").alignment = Alignment(horizontal="center")
ws5.cell(r, 6, "-").alignment = Alignment(horizontal="center")
for c in range(1, 7):
    ws5.cell(r, c).border = THIN_B
r += 1

ben_concessions = [
    ("Lower DSCR floor to 1.25x", "1.35x to 1.25x", ben_base_rate, ben_base_amort, 1.25, ben_base_ltv),
    ("Lower DSCR floor to 1.20x", "1.35x to 1.20x", ben_base_rate, ben_base_amort, 1.20, ben_base_ltv),
    ("Extend amort to 30 years", "25 to 30 yr", ben_base_rate, 30, ben_base_dscr, ben_base_ltv),
    ("Lower rate by 25 bps", "6.40% to 6.15%", 0.0615, ben_base_amort, ben_base_dscr, ben_base_ltv),
    ("Higher LTV cap to 65%", "60% to 65%", ben_base_rate, ben_base_amort, ben_base_dscr, 0.65),
    ("Higher LTV cap to 70%", "60% to 70%", ben_base_rate, ben_base_amort, ben_base_dscr, 0.70),
    ("Use 5.5% cap rate (vs 6%)", "Cap rate 6% to 5.5%", ben_base_rate, ben_base_amort, ben_base_dscr, ben_base_ltv),
    ("Combo: 1.25x DSCR + 65% LTV + 30-yr", "All concessions", ben_base_rate, 30, 1.25, 0.65),
]

for i, (name, change, rate, amort, dscr, ltv) in enumerate(ben_concessions):
    if "5.5% cap" in name:
        new_loan, _, _ = calc_max_loan(rate, amort, dscr, ltv, cap_rate=0.055)
    else:
        new_loan, _, _ = calc_max_loan(rate, amort, dscr, ltv)
    delta = new_loan - base_loan
    ws5.cell(r, 1, name).font = BODY
    ws5.cell(r, 2, change).font = BODY
    ws5.cell(r, 2).alignment = Alignment(horizontal="center")
    ws5.cell(r, 3, new_loan).number_format = MONEY
    ws5.cell(r, 4, delta).number_format = MONEY
    if delta > 0:
        ws5.cell(r, 4).fill = GREEN_FILL
        ws5.cell(r, 4).font = BOLD
    ws5.cell(r, 5, delta).number_format = MONEY
    if delta > 0:
        ws5.cell(r, 5).fill = GREEN_FILL
    pct_move = delta / base_loan if base_loan else 0
    ws5.cell(r, 6, pct_move).number_format = PCT0
    ws5.cell(r, 6).alignment = Alignment(horizontal="right")
    if delta > 0:
        ws5.cell(r, 6).fill = GREEN_FILL
    for c in range(1, 7):
        ws5.cell(r, c).border = THIN_B
    r += 1

r += 2

# Section: ServisFirst concessions
ws5.cell(r, 1, "SERVISFIRST CONCESSION ANALYSIS").font = SECTION_FONT
ws5.cell(r, 1).border = THICK_B
for c in range(2, 7):
    ws5.cell(r, c).border = THICK_B
r += 1

ws5.cell(r, 1, "Base case: 6.50% / 20-yr / 1.50x DSCR / 75% LTV").font = NOTE_FONT
ws5.merge_cells(f"A{r}:F{r}")
r += 2

# Header
ws5.cell(r, 1, "Concession").fill = DARK
ws5.cell(r, 1).font = HEADER_FONT
ws5.cell(r, 2, "Change").fill = DARK
ws5.cell(r, 2).font = HEADER_FONT
ws5.cell(r, 2).alignment = Alignment(horizontal="center")
ws5.cell(r, 3, "New Max Loan").fill = DARK
ws5.cell(r, 3).font = HEADER_FONT
ws5.cell(r, 3).alignment = Alignment(horizontal="center")
ws5.cell(r, 4, "vs Base").fill = DARK
ws5.cell(r, 4).font = HEADER_FONT
ws5.cell(r, 4).alignment = Alignment(horizontal="center")
ws5.cell(r, 5, "$ Value").fill = DARK
ws5.cell(r, 5).font = HEADER_FONT
ws5.cell(r, 5).alignment = Alignment(horizontal="center")
ws5.cell(r, 6, "% Move").fill = DARK
ws5.cell(r, 6).font = HEADER_FONT
ws5.cell(r, 6).alignment = Alignment(horizontal="center")
r += 1

sf_base_rate = 0.0650
sf_base_amort = 20
sf_base_dscr = 1.50
sf_base_ltv = 0.75

sf_base_loan, _, _ = calc_max_loan(sf_base_rate, sf_base_amort, sf_base_dscr, sf_base_ltv)
ws5.cell(r, 1, "BASE CASE").font = BOLD
ws5.cell(r, 1).fill = LIGHT_GOLD
ws5.cell(r, 2, "-").alignment = Alignment(horizontal="center")
ws5.cell(r, 3, sf_base_loan).number_format = MONEY
ws5.cell(r, 3).font = BOLD
ws5.cell(r, 3).fill = LIGHT_GOLD
ws5.cell(r, 4, 0).number_format = MONEY
ws5.cell(r, 5, "-").alignment = Alignment(horizontal="center")
ws5.cell(r, 6, "-").alignment = Alignment(horizontal="center")
for c in range(1, 7):
    ws5.cell(r, c).border = THIN_B
r += 1

sf_concessions = [
    ("Extend amort to 25 years", "20 to 25 yr", sf_base_rate, 25, sf_base_dscr, sf_base_ltv),
    ("Extend amort to 30 years", "20 to 30 yr", sf_base_rate, 30, sf_base_dscr, sf_base_ltv),
    ("Lower DSCR floor to 1.35x", "1.50x to 1.35x", sf_base_rate, sf_base_amort, 1.35, sf_base_ltv),
    ("Lower DSCR floor to 1.25x", "1.50x to 1.25x", sf_base_rate, sf_base_amort, 1.25, sf_base_ltv),
    ("Lower rate to 6.25%", "6.50% to 6.25%", 0.0625, sf_base_amort, sf_base_dscr, sf_base_ltv),
    ("Lower rate to 6.00%", "6.50% to 6.00%", 0.0600, sf_base_amort, sf_base_dscr, sf_base_ltv),
    ("Combo: 25-yr + 1.35x", "Two concessions", sf_base_rate, 25, 1.35, sf_base_ltv),
    ("Combo: 25-yr + 1.35x + 6.25%", "Three concessions", 0.0625, 25, 1.35, sf_base_ltv),
    ("Combo: 30-yr + 1.25x + 6.25%", "Match Ben Option 2 effectively", 0.0625, 30, 1.25, sf_base_ltv),
]

for name, change, rate, amort, dscr, ltv in sf_concessions:
    new_loan, _, _ = calc_max_loan(rate, amort, dscr, ltv)
    delta = new_loan - sf_base_loan
    ws5.cell(r, 1, name).font = BODY
    ws5.cell(r, 2, change).font = BODY
    ws5.cell(r, 2).alignment = Alignment(horizontal="center")
    ws5.cell(r, 3, new_loan).number_format = MONEY
    ws5.cell(r, 4, delta).number_format = MONEY
    if delta > 0:
        ws5.cell(r, 4).fill = GREEN_FILL
        ws5.cell(r, 4).font = BOLD
    ws5.cell(r, 5, delta).number_format = MONEY
    if delta > 0:
        ws5.cell(r, 5).fill = GREEN_FILL
    pct_move = delta / sf_base_loan if sf_base_loan else 0
    ws5.cell(r, 6, pct_move).number_format = PCT0
    ws5.cell(r, 6).alignment = Alignment(horizontal="right")
    if delta > 0:
        ws5.cell(r, 6).fill = GREEN_FILL
    for c in range(1, 7):
        ws5.cell(r, c).border = THIN_B
    r += 1

r += 2

# Strategic conclusions
ws5.cell(r, 1, "STRATEGIC CONCLUSIONS").font = SECTION_FONT
ws5.cell(r, 1).border = THICK_B
for c in range(2, 7):
    ws5.cell(r, c).border = THICK_B
r += 1

conclusions = [
    "1. AMORTIZATION extension is the highest-value single concession. Each 5 years of amort adds ~$200K-$300K of capacity.",
    "2. DSCR floor reduction is second-highest value. Moving from 1.50x to 1.25x adds ~$370K of capacity at ServisFirst's terms.",
    "3. Rate reductions are third-most valuable. Each 25 bps of rate reduction adds ~$50K-$80K of capacity.",
    "4. LTV ceiling matters less when DSCR is binding (which it usually is at conservative cap rates).",
    "5. For ServisFirst to genuinely beat Ben Opt 2, they need to give on amort (25-yr+) AND DSCR floor (1.35x or lower).",
    "6. For Ben to compete with a hypothetical aggressive ServisFirst offer, he can give on amort (30-yr) and/or LTV (65%).",
]
for c in conclusions:
    ws5.cell(r, 1, c).font = BODY
    ws5.merge_cells(f"A{r}:F{r}")
    r += 1

ws5.freeze_panes = "A5"


# ── Save ──
OUTPUT = Path(__file__).parent / "Casa_Yano_Loan_Evaluator.xlsx"
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"Tabs: {', '.join(wb.sheetnames)}")
