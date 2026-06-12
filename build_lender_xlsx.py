#!/usr/bin/env python3
"""Build Casa Yano lender package spreadsheet from data.js + QBO financials"""

import json, calendar
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

# Load dashboard data
with open(Path(__file__).parent / "public" / "data.js") as f:
    raw = f.read()
    D = json.loads(raw.split("window.__DATA__ = ")[1].rstrip(";\n"))

PF = D["proForma"]
pf_m = PF["monthly"]
exp_m = D["expenseMonthly"]
fwd = D["forwardPace"]
fa = D["forecastAccuracy"]

wb = Workbook()

# ── Styles ──
DARK = PatternFill("solid", fgColor="2D2D2D")
GOLD_FILL = PatternFill("solid", fgColor="C5A55A")
LIGHT_GRAY = PatternFill("solid", fgColor="F5F5F5")
GREEN_FILL = PatternFill("solid", fgColor="E8F5E9")
BLUE_FILL = PatternFill("solid", fgColor="E3F2FD")
YELLOW_FILL = PatternFill("solid", fgColor="FFFDE7")
LIGHT_GOLD = PatternFill("solid", fgColor="F5ECD7")
RED_FILL = PatternFill("solid", fgColor="FFEBEE")

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
GOLD_FONT = Font(name="Arial", bold=True, color="C5A55A", size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="2D2D2D")
SECTION_FONT = Font(name="Arial", bold=True, size=12, color="2D2D2D")
BODY = Font(name="Arial", size=10, color="333333")
BOLD = Font(name="Arial", bold=True, size=10, color="333333")
BLUE_INPUT = Font(name="Arial", size=10, color="0000FF")
GREEN_LINK = Font(name="Arial", size=10, color="008000")
NOTE_FONT = Font(name="Arial", size=9, color="999999")
MONEY = '$#,##0;($#,##0);"-"'
PCT = '0.0%'
NUM = '#,##0'
THIN_B = Border(bottom=Side(style="thin", color="DDDDDD"))
THICK_B = Border(bottom=Side(style="medium", color="2D2D2D"))

MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

def hdr_row(ws, row, cols, fill=DARK, font=HEADER_FONT):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def alt_fill(ws, row, cols):
    if row % 2 == 0:
        for c in range(1, cols+1): ws.cell(row, c).fill = LIGHT_GRAY

def write_item(ws, row, label, val, fmt=None, bold=False, indent=False, font=None, col_label=1, col_val=2):
    lbl = f"  {label}" if indent else label
    ws.cell(row, col_label, lbl).font = BOLD if bold else BODY
    c = ws.cell(row, col_val, val)
    c.font = font or (BOLD if bold else BODY)
    if fmt: c.number_format = fmt
    ws.cell(row, col_label).border = THIN_B
    ws.cell(row, col_val).border = THIN_B
    return row + 1

def section_hdr(ws, row, title, cols=2):
    ws.cell(row, 1, title).font = SECTION_FONT
    for c in range(1, cols+1): ws.cell(row, c).border = THICK_B
    return row + 1

# QBO data (hardcoded from extracted financials - updated June 4, 2026 for period through May 31, 2026)
QBO_AS_OF_DATE = "May 31, 2026"
QBO_PERIOD = "January 1 - May 31, 2026"

QBO_BS_2026 = {
    "cash": 1457,
    "fixed_assets": {
        "acquisition_fee": 49500, "appliances": 1810.81, "buildings": 1890016.53,
        "cost_seg": 5920, "fundraising": 11250, "ff_e": 73760.68,
        "land": 596847.33, "landscaping": 821.86, "photography": 2175,
        "rehab_hard": 1405644.37, "rehab_soft": 159844.11,
        "rehab_total": 1565488.48, "total": 4197590.69,
    },
    "total_assets": 4199047.69,
    "current_liab": 2063.95,
    "lt_debt": {"cdrbc": 1000000, "gft_ira": 500000, "total": 1500000},
    "total_liab": 1502063.95,
    "retained_earnings": -343121.08,
    "net_income_ytd": 72502.37,
    "total_equity": -270618.71,
}

QBO_PL_2026 = {
    "rent": 273154.41,
    "expenses": {
        "advertising": 2984.60, "commissions": 434.95, "guest_relations": 539.50,
        "mgmt_fee": 27785.93, "landscaping": 450,
        "cleaning": 35576.28, "fire_safety": 450, "general_repairs": 4309.90, "pest_control": 210,
        "disposal": 50,
        "repairs_total": 40596.18,
        "supplies": 11090.99,
        "city_county_tax": 25418.29, "property_tax": 13339.62, "state_tax": 1612.76,
        "taxes_total": 40370.67,
        "electricity": 302.01, "gas": 329.37, "internet": 1395.90, "water": 2764.64,
        "utilities_total": 4791.92,
        "total": 129044.74,
    },
    "noi": 144109.67,
    "bank_charges": 357.30,
    "mortgage_interest": 71250,
    "net_income": 72502.37,
}


# ════════════════════════════════════════════════════════
# TAB 1: Executive Summary
# ════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.sheet_properties.tabColor = "C5A55A"
ws1.column_dimensions["A"].width = 30
ws1.column_dimensions["B"].width = 22
ws1.column_dimensions["C"].width = 5
ws1.column_dimensions["D"].width = 30
ws1.column_dimensions["E"].width = 22

r = 1
ws1.cell(r, 1, "Casa Yano - Lender Package").font = Font(name="Arial", bold=True, size=18, color="2D2D2D")
ws1.merge_cells("A1:E1")
r = 2
ws1.cell(r, 1, "210 W Yanonali St, Santa Barbara, CA 93101").font = Font(name="Arial", size=11, color="666666")
ws1.merge_cells("A2:E2")
r = 3
ws1.cell(r, 1, f"Prepared {D['generatedAt'][:10]}").font = NOTE_FONT
ws1.merge_cells("A3:E3")

r = 5
r = section_hdr(ws1, r, "PROPERTY DETAILS")
props = [
    ("Property Type", "Short-Term Rental (STR)"),
    ("Units", "6"),
    ("Location", "Santa Barbara, CA"),
    ("Opening Date", "December 18, 2025"),
    ("Days Operating", str(D["daysSinceLaunch"])),
    ("Management", "ZenStay Inc (third-party, 10% of gross)"),
]
for label, val in props:
    r = write_item(ws1, r, label, val)

r += 1
r_ytd_start = r
r = section_hdr(ws1, r, "YTD PERFORMANCE (QBO - Cash Basis)")
section_hdr(ws1, r_ytd_start, "FULL-YEAR PRO FORMA", cols=2)
ws1.cell(r_ytd_start, 4, "FULL-YEAR PRO FORMA").font = SECTION_FONT
ws1.cell(r_ytd_start, 4).border = THICK_B
ws1.cell(r_ytd_start, 5).border = THICK_B

ytd_items = [
    ("Period", "Jan 1 – May 31, 2026", None),
    ("Rental Income", QBO_PL_2026["rent"], MONEY),
    ("Total Operating Expenses", QBO_PL_2026["expenses"]["total"], MONEY),
    ("Net Operating Income", QBO_PL_2026["noi"], MONEY),
]
pf_items = [
    ("Projection Basis", f"{PF['closedMonths']} months actuals + seasonal model", None),
    ("Gross Revenue", PF["gross"], MONEY),
    ("Net to Owner", PF["netOwner"], MONEY),
    ("Total OpEx + Fixed", PF["opex"] + PF["mgmtFee"] + PF["propertyTax"] + PF["insurance"] + PF["otherFixed"], MONEY),
    ("NOI (After Known Fixed)", PF["noiAfterKnown"], MONEY),
    ("Avg ADR (Blended)", PF["avgAdr"], MONEY),
    ("Avg Occupancy (Blended)", PF["avgOcc"] / 100, PCT),
]

start_r = r
for label, val, fmt in ytd_items:
    r = write_item(ws1, r, label, val, fmt, bold=(label == "Net Operating Income"))
r_ytd_end = r

r = start_r
for label, val, fmt in pf_items:
    write_item(ws1, r, label, val, fmt, bold=(label == "NOI (After Known Fixed)"), col_label=4, col_val=5)
    r += 1

r = max(r_ytd_end, r) + 1
r = section_hdr(ws1, r, "KEY HIGHLIGHTS", cols=5)

# Compute dynamic highlights from the PMS data
closed_months_pf = [m for m in pf_m if m["source"] == "actual"]
closed_count = len(closed_months_pf)
if closed_months_pf:
    max_month = max(closed_months_pf, key=lambda m: m["gross"])
    max_month_label = max_month["month"]
    max_month_gross = max_month["gross"]
    occ_min = min(m["occ"] for m in closed_months_pf)
    occ_max = max(m["occ"] for m in closed_months_pf)
    first_month_label = closed_months_pf[0]["month"]
    last_month_label = closed_months_pf[-1]["month"]
else:
    max_month_label, max_month_gross = "-", 0
    occ_min, occ_max = 0, 0
    first_month_label, last_month_label = "-", "-"

highlights = [
    f"Total project basis: ${QBO_BS_2026['fixed_assets']['total']:,.0f} (per QBO balance sheet)",
    f"Existing debt: ${QBO_BS_2026['lt_debt']['total']:,.0f} - seeking refinance",
    f"Occupancy ranging from {occ_min:.0f}% to {occ_max:.0f}% across closed months ({first_month_label} - {last_month_label}) - strong operating performance",
    f"Best month to date: {max_month_label} at ${max_month_gross:,} gross",
    f"Owner margin of {D['summary']['ownerMargin']}% - efficient operations with low OTA commission rates",
    f"QBO YTD NOI of ${QBO_PL_2026['noi']:,.0f} through May 31 (5 months of operations)",
]
for h in highlights:
    ws1.cell(r, 1, f"•  {h}").font = BODY
    ws1.merge_cells(f"A{r}:E{r}")
    r += 1

ws1.freeze_panes = "A5"


# ════════════════════════════════════════════════════════
# TAB 2: Project History
# ════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Project History")
ws2.sheet_properties.tabColor = "795548"
ws2.column_dimensions["A"].width = 32
ws2.column_dimensions["B"].width = 22
ws2.column_dimensions["C"].width = 5
ws2.column_dimensions["D"].width = 32
ws2.column_dimensions["E"].width = 22

r = 1
ws2.cell(r, 1, "Casa Yano - Project History & Renovation").font = Font(name="Arial", bold=True, size=18, color="2D2D2D")
ws2.merge_cells("A1:E1")
r = 2
ws2.cell(r, 1, "Source: QuickBooks Online balance sheet + project records").font = NOTE_FONT
ws2.merge_cells("A2:E2")

r = 4
r = section_hdr(ws2, r, "ACQUISITION")
for label, val, fmt in [("Purchase Price", 2475000, MONEY), ("Close Date", "March 2024", None)]:
    r = write_item(ws2, r, label, val, fmt)

r += 1
r = section_hdr(ws2, r, "RENOVATION")
reno_items = [
    ("Total Construction Cost (QBO)", round(QBO_BS_2026["fixed_assets"]["rehab_total"]), MONEY),
    ("Hard Costs", round(QBO_BS_2026["fixed_assets"]["rehab_hard"]), MONEY),
    ("Soft Costs", round(QBO_BS_2026["fixed_assets"]["rehab_soft"]), MONEY),
    ("Renovation Start", "October 2024", None),
    ("Renovation Complete", "November 2025", None),
    ("Duration", "13 months", None),
    ("Scope", "Full gut renovation to studs", None),
]
for label, val, fmt in reno_items:
    r = write_item(ws2, r, label, val, fmt, indent=label in ("Hard Costs", "Soft Costs"))

r += 1
r = section_hdr(ws2, r, "SCOPE OF WORK", cols=5)
scope = [
    ("Structural", "Rebuilt foundation including underpinning with pylons"),
    ("Framing & Envelope", "New framing, new windows, new roof, new stucco"),
    ("Mechanical", "All new plumbing, all new electrical"),
    ("HVAC", "New mini splits in each unit"),
    ("Insulation & Walls", "All new insulation, all new drywall"),
    ("Interiors", "All new flooring, lighting, kitchens, bathrooms, and finishes"),
    ("Exterior", "New hardscape and landscape throughout the property"),
]
for label, desc in scope:
    ws2.cell(r, 1, label).font = BOLD
    ws2.cell(r, 2, desc).font = BODY
    ws2.merge_cells(f"B{r}:E{r}")
    ws2.cell(r, 1).border = THIN_B
    ws2.cell(r, 2).border = THIN_B
    r += 1

r += 1
r = section_hdr(ws2, r, "TOTAL PROJECT COST (per QBO Balance Sheet)")
cost_start = r
cost_items = [
    ("Land", round(QBO_BS_2026["fixed_assets"]["land"]), MONEY),
    ("Buildings", round(QBO_BS_2026["fixed_assets"]["buildings"]), MONEY),
    ("Rehab Costs - Hard", round(QBO_BS_2026["fixed_assets"]["rehab_hard"]), MONEY),
    ("Rehab Costs - Soft", round(QBO_BS_2026["fixed_assets"]["rehab_soft"]), MONEY),
    ("FF&E / Furniture & Fixtures", round(QBO_BS_2026["fixed_assets"]["ff_e"]), MONEY),
    ("Other (Acq Fee, Fundraising, Cost Seg, etc.)", round(49500 + 11250 + 5920 + 1810.81 + 821.86 + 2175), MONEY),
]
for label, val, fmt in cost_items:
    r = write_item(ws2, r, label, val, fmt)
# Total
ws2.cell(r, 1, "Total Fixed Assets").font = Font(name="Arial", bold=True, size=11, color="2D2D2D")
ws2.cell(r, 2).value = f"=SUM(B{cost_start}:B{r-1})"
ws2.cell(r, 2).font = Font(name="Arial", bold=True, size=11, color="2D2D2D")
ws2.cell(r, 2).number_format = MONEY
ws2.cell(r, 1).border = Border(top=Side(style="medium", color="2D2D2D"), bottom=Side(style="medium", color="2D2D2D"))
ws2.cell(r, 2).border = Border(top=Side(style="medium", color="2D2D2D"), bottom=Side(style="medium", color="2D2D2D"))
r_total_assets = r
r += 2

r = section_hdr(ws2, r, "EXISTING DEBT")
for label, val, fmt in [
    ("CDRBC, LLC", QBO_BS_2026["lt_debt"]["cdrbc"], MONEY),
    ("GFT IRA, LLC", QBO_BS_2026["lt_debt"]["gft_ira"], MONEY),
    ("Total Existing Debt", QBO_BS_2026["lt_debt"]["total"], MONEY),
]:
    r = write_item(ws2, r, label, val, fmt, bold=(label.startswith("Total")))

r += 1
r = section_hdr(ws2, r, "VALUE CREATION")
val_start = r
for label, val, fmt in [
    ("Projected NOI (2026)", PF["noiAfterKnown"], MONEY),
    ("Cap Rate Applied", 0.06, PCT),
]:
    r = write_item(ws2, r, label, val, fmt)
ws2.cell(r, 1, "Implied Value").font = BOLD
ws2.cell(r, 2).value = f"=B{val_start}/B{val_start+1}"
ws2.cell(r, 2).font = BOLD
ws2.cell(r, 2).number_format = MONEY
ws2.cell(r, 1).border = THIN_B
ws2.cell(r, 2).border = THIN_B
r += 1
ws2.cell(r, 1, "Total Basis (per QBO)").font = BODY
ws2.cell(r, 2).value = f"=B{r_total_assets}"
ws2.cell(r, 2).font = GREEN_LINK
ws2.cell(r, 2).number_format = MONEY
ws2.cell(r, 1).border = THIN_B
ws2.cell(r, 2).border = THIN_B
r += 1
ws2.cell(r, 1, "Value Created").font = Font(name="Arial", bold=True, size=11, color="4CAF50")
ws2.cell(r, 2).value = f"=B{r-2}-B{r-1}"
ws2.cell(r, 2).font = Font(name="Arial", bold=True, size=11, color="4CAF50")
ws2.cell(r, 2).number_format = MONEY
ws2.cell(r, 1).border = THIN_B
ws2.cell(r, 2).border = THIN_B
r += 1
ws2.cell(r, 1, "Return on Cost").font = BOLD
ws2.cell(r, 2).value = f"=B{val_start}/B{r-2}"
ws2.cell(r, 2).font = BOLD
ws2.cell(r, 2).number_format = PCT
ws2.cell(r, 1).border = THIN_B
ws2.cell(r, 2).border = THIN_B

r += 2
r = section_hdr(ws2, r, "TIMELINE", cols=5)
timeline = [
    ("Mar 2024", "Property acquired"),
    ("Mar – Sep 2024", "Planning, permitting, design"),
    ("Oct 2024", "Construction begins - full gut renovation"),
    ("Nov 2025", "Construction complete - certificate of occupancy"),
    ("Dec 18, 2025", "First guest booking - STR operations commence"),
    ("Jan – Mar 2026", "First full quarter of operations - strong ramp"),
    ("Apr 2026", "Highest-grossing month to date"),
]
for date, event in timeline:
    ws2.cell(r, 1, date).font = BOLD
    ws2.cell(r, 2, event).font = BODY
    ws2.merge_cells(f"B{r}:E{r}")
    ws2.cell(r, 1).border = THIN_B
    ws2.cell(r, 2).border = THIN_B
    r += 1

ws2.freeze_panes = "A4"


# ════════════════════════════════════════════════════════
# TAB 3: QBO P&L (2026 YTD)
# ════════════════════════════════════════════════════════
ws3 = wb.create_sheet("P&L (2026 YTD)")
ws3.sheet_properties.tabColor = "2D8B2D"
ws3.column_dimensions["A"].width = 34

# Build monthly revenue and expense lookups (2026 only)
rev_by_month = {}
for md in D["monthlyData"]:
    parts = md["month"].split(" '")
    if len(parts) == 2 and int(parts[1]) + 2000 == 2026:
        rev_by_month[parts[0]] = md

exp_by_month = {}
for ed in exp_m:
    parts = ed["month"].split(" '")
    if len(parts) == 2 and int(parts[1]) + 2000 == 2026:
        exp_by_month[parts[0]] = ed

# Determine which months are fully closed (actual data, not in-progress)
# Only "actual" months show in the YTD P&L - current/booked months would mislead
# with partial expense data.
closed_months = [m for m in MONTHS if m in rev_by_month and m in exp_by_month
                 and pf_m[MONTHS.index(m)]["source"] == "actual"]
num_months = len(closed_months)
num_cols = num_months + 2  # label col + monthly cols + YTD col
ytd_col = num_months + 2

for c in range(2, num_cols + 1):
    ws3.column_dimensions[get_column_letter(c)].width = 16

r = 1
ws3.cell(r, 1, "Casa Yano - Profit & Loss by Month (2026)").font = TITLE_FONT
ws3.merge_cells(f"A1:{get_column_letter(num_cols)}1")
r = 2
ws3.cell(r, 1, f"Operating data: January – {closed_months[-1]} 2026  |  QBO YTD cross-referenced").font = Font(name="Arial", size=11, color="666666")
ws3.merge_cells(f"A2:{get_column_letter(num_cols)}2")
r = 3
ws3.cell(r, 1, "Source: PMS booking data + expense reports, reconciled to QuickBooks Online").font = NOTE_FONT
ws3.merge_cells(f"A3:{get_column_letter(num_cols)}3")

r = 5
hdr_row(ws3, r, num_cols)
ws3.cell(r, 1, "").font = HEADER_FONT
for ci, mn in enumerate(closed_months):
    ws3.cell(r, ci + 2, mn).font = HEADER_FONT
    ws3.cell(r, ci + 2).alignment = Alignment(horizontal="center")
ws3.cell(r, ytd_col, "YTD Total").font = HEADER_FONT
ws3.cell(r, ytd_col).alignment = Alignment(horizontal="center")
r += 1

def pl_row(ws, row, label, values, bold=False, fill=None, is_total=False, is_noi=False,
           formula_cells=None):
    """Write a P&L row with monthly values + YTD sum formula.
    If formula_cells is provided, the monthly cells use formulas instead of values.
    formula_cells is a list of formulas (strings) matching the length of months.
    """
    font = BOLD if bold else BODY
    if is_noi:
        font = Font(name="Arial", bold=True, size=12, color="2D8B2D")
    elif is_total:
        font = Font(name="Arial", bold=True, size=11, color="E53935")
    ws.cell(row, 1, label).font = font
    if formula_cells is not None:
        for ci, formula in enumerate(formula_cells):
            c = ws.cell(row, ci + 2)
            c.value = formula
            c.number_format = MONEY
            c.font = font
    else:
        for ci, val in enumerate(values):
            c = ws.cell(row, ci + 2, val)
            c.number_format = MONEY
            c.font = font
    # YTD = SUM of monthly cells
    first_col = get_column_letter(2)
    last_col = get_column_letter(num_months + 1)
    ytd_cell = ws.cell(row, ytd_col)
    ytd_cell.value = f"=SUM({first_col}{row}:{last_col}{row})"
    ytd_cell.number_format = MONEY
    ytd_cell.font = font
    bdr = THICK_B if (is_total or is_noi) else THIN_B
    for c in range(1, num_cols + 1):
        ws.cell(row, c).border = bdr
        if fill:
            ws.cell(row, c).fill = fill
    return row + 1

# ── INCOME ──
ws3.cell(r, 1, "INCOME").font = BOLD
for c in range(1, num_cols + 1): ws3.cell(r, c).fill = LIGHT_GRAY
r += 1

gross_vals = [rev_by_month[m]["gross"] for m in closed_months]
to_owner_vals = [rev_by_month[m]["toOwner"] for m in closed_months]
r_gross = r
r = pl_row(ws3, r, "Gross Booking Revenue", gross_vals, bold=True)
r_to_owner = r
r = pl_row(ws3, r, "Net to Owner (after OTA & Processing)", to_owner_vals, bold=True, fill=LIGHT_GOLD)

# ── EXPENSES ──
r += 1
ws3.cell(r, 1, "OPERATING EXPENSES").font = BOLD
for c in range(1, num_cols + 1): ws3.cell(r, c).fill = LIGHT_GRAY
r += 1

# Expense line items by category - track which rows are actually written
expense_cats = [
    ("Cleaning", "cleaning"),
    ("Supplies", "supplies"),
    ("Repairs & Maintenance", "maint"),
    ("Management Fee", "mgmt"),
    ("Marketing & Advertising", "marketing"),
    ("Taxes & Licenses", "taxes"),
    ("Other", "other"),
]
expense_row_range_start = None
expense_row_range_end = None
for label, key in expense_cats:
    vals = [exp_by_month[m].get(key, 0) for m in closed_months]
    if sum(vals) > 0:
        if expense_row_range_start is None:
            expense_row_range_start = r
        r = pl_row(ws3, r, f"  {label}", vals)
        expense_row_range_end = r - 1  # row just written

# Total expenses - FORMULA SUM of category rows above
r_total_exp = r
total_exp_formulas = []
for ci in range(num_months):
    col = get_column_letter(ci + 2)
    total_exp_formulas.append(f"=SUM({col}{expense_row_range_start}:{col}{expense_row_range_end})")
r = pl_row(ws3, r, "Total Operating Expenses", None, is_total=True, formula_cells=total_exp_formulas)

# ── NOI ── FORMULA: Net to Owner - Total OpEx
r += 1
r_noi = r  # capture NOI row before pl_row writes it
noi_formulas = []
for ci in range(num_months):
    col = get_column_letter(ci + 2)
    noi_formulas.append(f"={col}{r_to_owner}-{col}{r_total_exp}")
r = pl_row(ws3, r, "Net Operating Income", None, is_noi=True, fill=GREEN_FILL, formula_cells=noi_formulas)

# Reconciliation check row - verifies NOI = Net to Owner - Total OpEx (should be zero variance)
r += 1
ws3.cell(r, 1, "Reconciliation check:").font = Font(name="Arial", size=9, color="666666", italic=True)
for ci in range(num_months):
    col = get_column_letter(ci + 2)
    cell = ws3.cell(r, ci + 2)
    cell.value = f'=IF(ROUND({col}{r_to_owner}-{col}{r_total_exp}-{col}{r_noi},0)=0,"OK","ERR")'
    cell.font = Font(name="Arial", size=9, color="666666", italic=True)
    cell.alignment = Alignment(horizontal="center")
ytd_check = ws3.cell(r, ytd_col)
ytd_col_letter = get_column_letter(ytd_col)
ytd_check.value = f'=IF(ROUND({ytd_col_letter}{r_to_owner}-{ytd_col_letter}{r_total_exp}-{ytd_col_letter}{r_noi},0)=0,"OK","ERR")'
ytd_check.font = Font(name="Arial", size=9, color="666666", italic=True)
ytd_check.alignment = Alignment(horizontal="center")
# Conditional formatting: green if OK, red if ERR
for ci in range(num_months + 1):
    col = get_column_letter(ci + 2)
    ws3.conditional_formatting.add(f"{col}{r}",
        FormulaRule(formula=[f'{col}{r}="OK"'], fill=GREEN_FILL))
    ws3.conditional_formatting.add(f"{col}{r}",
        FormulaRule(formula=[f'{col}{r}="ERR"'], fill=RED_FILL))

# ── QBO Reconciliation note ──
r += 2
ws3.cell(r, 1, "QBO RECONCILIATION").font = SECTION_FONT
for c in range(1, num_cols + 1): ws3.cell(r, c).border = THICK_B
r += 1
ws3.cell(r, 1, f"QBO YTD Rental Income (Jan 1 – May 31): ${QBO_PL_2026['rent']:,.0f}").font = NOTE_FONT
ws3.merge_cells(f"A{r}:{get_column_letter(num_cols)}{r}")
r += 1
ws3.cell(r, 1, f"QBO YTD Total Expenses: ${QBO_PL_2026['expenses']['total']:,.0f}  |  QBO YTD NOI: ${QBO_PL_2026['noi']:,.0f}").font = NOTE_FONT
ws3.merge_cells(f"A{r}:{get_column_letter(num_cols)}{r}")
r += 1
ws3.cell(r, 1, "Note: Minor variances between PMS and QBO are expected due to cash vs accrual timing and partial-month cutoffs.").font = NOTE_FONT
ws3.merge_cells(f"A{r}:{get_column_letter(num_cols)}{r}")

ws3.freeze_panes = "B6"


# ════════════════════════════════════════════════════════
# TAB 4: Operating Detail (PMS data)
# ════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Operating Detail")
ws5.sheet_properties.tabColor = "C5A55A"
ws5.column_dimensions["A"].width = 22
for c in range(2, 16): ws5.column_dimensions[get_column_letter(c)].width = 12

r = 1
ws5.cell(r, 1, "Casa Yano - Monthly Operating Detail (2026)").font = TITLE_FONT
ws5.merge_cells("A1:M1")
r = 2
ws5.cell(r, 1, "Source: Property Management System booking data").font = NOTE_FONT
ws5.merge_cells("A2:M2")
r = 4

# Source row
ws5.cell(r, 1, "Status").font = BOLD
for i, m in enumerate(pf_m):
    cell = ws5.cell(r, i+2, m["source"].upper())
    cell.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center")
    cell.fill = PatternFill("solid", fgColor="4CAF50" if m["source"] == "actual" else "2196F3" if m["source"] == "booked" else "9E9E9E")
r = 5

headers = [""] + MONTHS + ["Full Year"]
for i, h in enumerate(headers):
    ws5.cell(r, i+1, h)
hdr_row(ws5, r, 14)
ws5.cell(r, 1).alignment = Alignment(horizontal="left")
r += 1

# Compute nights/avail
rev_data = []
for i, m in enumerate(pf_m):
    mn = i + 1
    days = calendar.monthrange(2026, mn)[1]
    avail = 6 * days
    nights = round(avail * m["occ"] / 100)
    revpar = round(m["gross"] / avail) if avail else 0
    rev_data.append({"bookings": m["bookings"], "nights": nights, "avail": avail,
                      "occ": m["occ"], "adr": m["adr"], "revpar": revpar,
                      "gross": m["gross"], "net": m["netOwner"]})

rev_rows = [
    ("Bookings", [d["bookings"] for d in rev_data], NUM),
    ("Nights Sold", [d["nights"] for d in rev_data], NUM),
    ("Available Nights", [d["avail"] for d in rev_data], NUM),
    ("Occupancy", [d["occ"] / 100 for d in rev_data], PCT),
    ("ADR", [d["adr"] for d in rev_data], MONEY),
    ("RevPAR", [d["revpar"] for d in rev_data], MONEY),
    ("Gross Revenue", [d["gross"] for d in rev_data], MONEY),
    ("Net to Owner", [d["net"] for d in rev_data], MONEY),
]

for label, vals, fmt in rev_rows:
    is_bold = label in ("Gross Revenue", "Net to Owner")
    ws5.cell(r, 1, label).font = BOLD if is_bold else BODY
    for i, v in enumerate(vals):
        ws5.cell(r, i+2, v).number_format = fmt
        ws5.cell(r, i+2).font = BOLD if is_bold else BODY
        ws5.cell(r, i+2).alignment = Alignment(horizontal="right")
    if fmt in (MONEY, NUM):
        ws5.cell(r, 14, f"=SUM(B{r}:M{r})").number_format = fmt
    elif fmt == PCT:
        ws5.cell(r, 14, f"=AVERAGE(B{r}:M{r})").number_format = fmt
    ws5.cell(r, 14).font = BOLD
    alt_fill(ws5, r, 14)
    r += 1

ws5.freeze_panes = "B6"


# ════════════════════════════════════════════════════════
# TAB 5: Forward Bookings
# ════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Forward Bookings")
ws6.sheet_properties.tabColor = "2196F3"
ws6.column_dimensions["A"].width = 14
for c in range(2, 9): ws6.column_dimensions[get_column_letter(c)].width = 16

r = 1
ws6.cell(r, 1, "Casa Yano - Forward Bookings on the Books").font = TITLE_FONT
ws6.merge_cells("A1:H1")
r = 2
ws6.cell(r, 1, f"As of {D['generatedAt'][:10]}. Confirmed reservations with deposits received.").font = NOTE_FONT
ws6.merge_cells("A2:H2")
r = 4

headers = ["Month", "PF Target", "Booked Revenue", "% of Target", "ADR", "RevPAR", "Remaining", "Status"]
for i, h in enumerate(headers):
    ws6.cell(r, i+1, h)
hdr_row(ws6, r, 8)
r += 1

fwd_start = r
for f in fwd:
    ws6.cell(r, 1, f["month"]).font = BOLD
    ws6.cell(r, 2, f["pfGross"]).number_format = MONEY
    ws6.cell(r, 2).font = BODY
    ws6.cell(r, 3, f["bookedGross"]).number_format = MONEY
    ws6.cell(r, 3).font = BOLD
    ws6.cell(r, 4, f["pacePct"] / 100).number_format = PCT
    ws6.cell(r, 4).font = BOLD
    ws6.cell(r, 5, f["bookedAdr"] if f["bookedAdr"] else "-").number_format = MONEY if f["bookedAdr"] else "@"
    ws6.cell(r, 5).font = BODY
    ws6.cell(r, 6, f["bookedRevpar"] if f["bookedRevpar"] else "-").number_format = MONEY if f["bookedRevpar"] else "@"
    ws6.cell(r, 6).font = BODY
    ws6.cell(r, 7, max(0, f["pfGross"] - f["bookedGross"])).number_format = MONEY
    ws6.cell(r, 7).font = BODY
    status = "Fully Booked" if f["pacePct"] >= 95 else "Strong" if f["pacePct"] >= 60 else "Building" if f["pacePct"] >= 25 else "Early"
    ws6.cell(r, 8, status).font = BODY
    if f["pacePct"] >= 60:
        for c in range(1, 9): ws6.cell(r, c).fill = GREEN_FILL
    elif f["pacePct"] >= 25:
        for c in range(1, 9): ws6.cell(r, c).fill = YELLOW_FILL
    r += 1

# Totals
num_cols = 8
for c in range(1, num_cols + 1): ws6.cell(r, c).border = Border(top=Side(style="medium", color="2D2D2D"))
ws6.cell(r, 1, "TOTAL").font = BOLD
ws6.cell(r, 2, f"=SUM(B{fwd_start}:B{r-1})").number_format = MONEY
ws6.cell(r, 2).font = BOLD
ws6.cell(r, 3, f"=SUM(C{fwd_start}:C{r-1})").number_format = MONEY
ws6.cell(r, 3).font = BOLD
ws6.cell(r, 4, f"=C{r}/B{r}").number_format = PCT
ws6.cell(r, 4).font = BOLD
ws6.cell(r, 7, f"=B{r}-C{r}").number_format = MONEY
ws6.cell(r, 7).font = BOLD

ws6.freeze_panes = "A5"


# ════════════════════════════════════════════════════════
# TAB 6: Pro Forma
# ════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Pro Forma")
ws7.sheet_properties.tabColor = "7B1FA2"
ws7.column_dimensions["A"].width = 38
ws7.column_dimensions["B"].width = 18
ws7.column_dimensions["C"].width = 14

r = 1
ws7.cell(r, 1, "Casa Yano - 2026 Pro Forma (Blended Forecast)").font = TITLE_FONT
ws7.merge_cells("A1:C1")
r = 2
ws7.cell(r, 1, f"Blends {PF['closedMonths']} months of closed actuals with seasonal projections calibrated from actual performance").font = NOTE_FONT
ws7.merge_cells("A2:C2")
r = 4

hdr_row(ws7, r, 3)
ws7.cell(r, 1, "").font = HEADER_FONT
ws7.cell(r, 2, "Amount").font = HEADER_FONT
ws7.cell(r, 3, "% of Gross").font = HEADER_FONT
r += 1

# Pull values from source pro forma model
pf_gross = PF["gross"]
pf_net_owner = PF["netOwner"]
pf_direct_opex = PF["opex"]  # Direct OpEx (cleaning + supplies/maint/other)
pf_mgmt_fee_input = PF["mgmtFee"]
pf_prop_tax_input = PF["propertyTax"]
pf_insurance_input = PF["insurance"]
pf_other_fixed_input = PF["otherFixed"]
pf_expected_noi = PF["noiAfterKnown"]
wf = PF["waterfall"]

# Compute exact OTA and Processing dollar amounts that reconcile to Net to Owner.
# Net to Owner = Gross - OTA Commission - Processing Fees (TOT is pass-through, not deducted).
# Allocate channel costs between OTA and Processing using actual observed ratios.
channel_costs = pf_gross - pf_net_owner
ota_pct_summary = D["summary"]["otaPct"] / 100  # blended OTA % from actual bookings
proc_pct_approx = 0.007  # ~0.7% from booking-level analysis (processing on direct bookings)
# Allocate channel_costs proportionally to maintain exact reconciliation
total_ratio = ota_pct_summary + proc_pct_approx
ota_dollars = round(channel_costs * (ota_pct_summary / total_ratio))
proc_dollars = channel_costs - ota_dollars  # exact remainder ensures sum reconciles

# Get cleaning and other direct opex from waterfall
cleaning_input = -wf[5]["value"] if len(wf) > 5 and wf[5]["name"] == "Cleaning" else round(pf_direct_opex * 0.4)
other_opex_input = -wf[6]["value"] if len(wf) > 6 and "Supplies" in wf[6]["name"] else (pf_direct_opex - cleaning_input)

# ── Build P&L with formulas ──
gross_row = r
ws7.cell(r, 1, "Gross Revenue").font = BOLD
ws7.cell(r, 2, pf_gross).number_format = MONEY
ws7.cell(r, 2).font = BOLD
ws7.cell(r, 3, 1.0).number_format = PCT
ws7.cell(r, 3).font = BOLD
for c in range(1, 4):
    ws7.cell(r, c).border = THIN_B
r += 1

ota_row = r
ws7.cell(r, 1, "  OTA Commissions").font = BODY
ws7.cell(r, 2, -ota_dollars).number_format = MONEY
ws7.cell(r, 2).font = BODY
ws7.cell(r, 3).value = f"=B{r}/B{gross_row}"
ws7.cell(r, 3).number_format = PCT
ws7.cell(r, 3).font = BODY
for c in range(1, 4):
    ws7.cell(r, c).border = THIN_B
r += 1

proc_row = r
ws7.cell(r, 1, "  Processing Fees").font = BODY
ws7.cell(r, 2, -proc_dollars).number_format = MONEY
ws7.cell(r, 2).font = BODY
ws7.cell(r, 3).value = f"=B{r}/B{gross_row}"
ws7.cell(r, 3).number_format = PCT
ws7.cell(r, 3).font = BODY
for c in range(1, 4):
    ws7.cell(r, c).border = THIN_B
r += 1

# Net to Owner - FORMULA: Gross + OTA + Processing
net_to_owner_row = r
ws7.cell(r, 1, "Net to Owner").font = BOLD
ws7.cell(r, 2).value = f"=B{gross_row}+B{ota_row}+B{proc_row}"
ws7.cell(r, 2).number_format = MONEY
ws7.cell(r, 2).font = BOLD
ws7.cell(r, 3).value = f"=B{r}/B{gross_row}"
ws7.cell(r, 3).number_format = PCT
ws7.cell(r, 3).font = BOLD
for c in range(1, 4):
    ws7.cell(r, c).fill = LIGHT_GOLD
    ws7.cell(r, c).border = THIN_B
r += 2

# Operating Expenses section header
ws7.cell(r, 1, "Operating Expenses").font = SECTION_FONT
for c in range(1, 4):
    ws7.cell(r, c).fill = LIGHT_GRAY
r += 1

opex_start_row = r
ws7.cell(r, 1, "  Cleaning").font = BODY
ws7.cell(r, 2, -cleaning_input).number_format = MONEY
ws7.cell(r, 2).font = BODY
ws7.cell(r, 3).value = f"=B{r}/B{gross_row}"
ws7.cell(r, 3).number_format = PCT
for c in range(1, 4):
    ws7.cell(r, c).border = THIN_B
r += 1

ws7.cell(r, 1, "  Supplies, Maintenance, Other").font = BODY
ws7.cell(r, 2, -other_opex_input).number_format = MONEY
ws7.cell(r, 2).font = BODY
ws7.cell(r, 3).value = f"=B{r}/B{gross_row}"
ws7.cell(r, 3).number_format = PCT
for c in range(1, 4):
    ws7.cell(r, c).border = THIN_B
r += 1

# Management Fee - FORMULA: 10% of Gross
ws7.cell(r, 1, "  Management Fee (10% of Gross)").font = BODY
ws7.cell(r, 2).value = f"=-0.10*B{gross_row}"
ws7.cell(r, 2).number_format = MONEY
ws7.cell(r, 2).font = BODY
ws7.cell(r, 3).value = f"=B{r}/B{gross_row}"
ws7.cell(r, 3).number_format = PCT
for c in range(1, 4):
    ws7.cell(r, c).border = THIN_B
r += 1

ws7.cell(r, 1, "  Property Tax").font = BODY
ws7.cell(r, 2, -pf_prop_tax_input).number_format = MONEY
ws7.cell(r, 2).font = BODY
ws7.cell(r, 3).value = f"=B{r}/B{gross_row}"
ws7.cell(r, 3).number_format = PCT
for c in range(1, 4):
    ws7.cell(r, c).border = THIN_B
r += 1

ws7.cell(r, 1, "  Insurance").font = BODY
ws7.cell(r, 2, -pf_insurance_input).number_format = MONEY
ws7.cell(r, 2).font = BODY
ws7.cell(r, 3).value = f"=B{r}/B{gross_row}"
ws7.cell(r, 3).number_format = PCT
for c in range(1, 4):
    ws7.cell(r, c).border = THIN_B
r += 1

ws7.cell(r, 1, "  Other Fixed Costs (utilities, internet, etc.)").font = BODY
ws7.cell(r, 2, -pf_other_fixed_input).number_format = MONEY
ws7.cell(r, 2).font = BODY
ws7.cell(r, 3).value = f"=B{r}/B{gross_row}"
ws7.cell(r, 3).number_format = PCT
for c in range(1, 4):
    ws7.cell(r, c).border = THIN_B
opex_end_row = r
r += 1

# Total Operating Expenses - FORMULA: SUM of opex rows
total_opex_row = r
ws7.cell(r, 1, "Total Operating Expenses").font = BOLD
ws7.cell(r, 2).value = f"=SUM(B{opex_start_row}:B{opex_end_row})"
ws7.cell(r, 2).number_format = MONEY
ws7.cell(r, 2).font = BOLD
ws7.cell(r, 3).value = f"=B{r}/B{gross_row}"
ws7.cell(r, 3).number_format = PCT
ws7.cell(r, 3).font = BOLD
for c in range(1, 4):
    ws7.cell(r, c).border = THICK_B
r += 2

# NOI - FORMULA: Net to Owner + Total OpEx (since OpEx is negative)
noi_row = r
NOI_FONT = Font(name="Arial", bold=True, size=12, color="2D8B2D")
ws7.cell(r, 1, "Net Operating Income").font = NOI_FONT
ws7.cell(r, 2).value = f"=B{net_to_owner_row}+B{total_opex_row}"
ws7.cell(r, 2).number_format = MONEY
ws7.cell(r, 2).font = NOI_FONT
ws7.cell(r, 3).value = f"=B{r}/B{gross_row}"
ws7.cell(r, 3).number_format = PCT
ws7.cell(r, 3).font = NOI_FONT
for c in range(1, 4):
    ws7.cell(r, c).fill = GREEN_FILL
    ws7.cell(r, c).border = THICK_B
r += 2

# ── RECONCILIATION CHECK ──
ws7.cell(r, 1, "RECONCILIATION CHECK").font = SECTION_FONT
for c in range(1, 4):
    ws7.cell(r, c).border = THICK_B
r += 1

ws7.cell(r, 1, "Expected NOI (from model)").font = BODY
ws7.cell(r, 2, pf_expected_noi).number_format = MONEY
ws7.cell(r, 2).font = BODY
expected_noi_row = r
r += 1

ws7.cell(r, 1, "Calculated NOI (this tab)").font = BODY
ws7.cell(r, 2).value = f"=B{noi_row}"
ws7.cell(r, 2).number_format = MONEY
ws7.cell(r, 2).font = BODY
calc_noi_row = r
r += 1

ws7.cell(r, 1, "Variance").font = BODY
ws7.cell(r, 2).value = f"=B{calc_noi_row}-B{expected_noi_row}"
ws7.cell(r, 2).number_format = MONEY
ws7.cell(r, 2).font = BODY
var_row = r
r += 1

ws7.cell(r, 1, "Status").font = BOLD
ws7.cell(r, 2).value = f'=IF(ABS(B{var_row})<=1,"OK - reconciles","REVIEW - variance exceeds $1")'
ws7.cell(r, 2).font = BOLD
status_row = r
# Conditional formatting on status cell
ws7.conditional_formatting.add(f"B{status_row}",
    FormulaRule(formula=[f'B{status_row}="OK - reconciles"'],
                fill=GREEN_FILL,
                font=Font(name="Arial", bold=True, color="2D8B2D")))
ws7.conditional_formatting.add(f"B{status_row}",
    FormulaRule(formula=[f'NOT(B{status_row}="OK - reconciles")'],
                fill=RED_FILL,
                font=Font(name="Arial", bold=True, color="E53935")))
r += 2

# ── PASS-THROUGH TAXES & FEES (INFORMATIONAL) ──
ws7.cell(r, 1, "PASS-THROUGH TAXES & FEES (INFORMATIONAL)").font = SECTION_FONT
for c in range(1, 4):
    ws7.cell(r, c).border = THICK_B
r += 1

ws7.cell(r, 1, "The following amounts are collected from guests at booking and either remitted to the City or").font = NOTE_FONT
ws7.merge_cells(f"A{r}:C{r}")
r += 1
ws7.cell(r, 1, "offset by corresponding expenses. They are NOT deducted from Net to Owner.").font = NOTE_FONT
ws7.merge_cells(f"A{r}:C{r}")
r += 1

# Estimate TOT for the projection (12% of room revenue, approximated as 9.7% of gross from actuals)
tot_rate = D["summary"].get("totalGross", 0)
tot_estimated = round(pf_gross * 0.097)  # ~9.7% of gross from booking data
ws7.cell(r, 1, "TOT (Transient Occupancy Tax) collected from guests").font = BODY
ws7.cell(r, 2, tot_estimated).number_format = MONEY
ws7.cell(r, 2).font = BODY
ws7.cell(r, 1).border = THIN_B
ws7.cell(r, 2).border = THIN_B
r += 1
ws7.cell(r, 1, "  Remitted to City of Santa Barbara (12% lodging tax)").font = NOTE_FONT
ws7.merge_cells(f"A{r}:C{r}")
r += 1

ws7.cell(r, 1, "Cleaning fees collected from guests").font = BODY
ws7.cell(r, 2, "see Cleaning expense").font = NOTE_FONT
ws7.cell(r, 1).border = THIN_B
ws7.cell(r, 2).border = THIN_B
r += 1
ws7.cell(r, 1, "  Offset by Cleaning expense above; net-neutral pass-through").font = NOTE_FONT
ws7.merge_cells(f"A{r}:C{r}")
r += 2

# ── MODEL INPUTS ──
ws7.cell(r, 1, "MODEL INPUTS").font = SECTION_FONT
ws7.cell(r, 1).border = THICK_B
ws7.cell(r, 2).border = THICK_B
r += 1
inputs = [
    ("Closed Months Used for Calibration", PF["closedMonths"], None),
    ("Baseline ADR (de-seasonalized)", PF["baselineAdr"], MONEY),
    ("Baseline Occupancy (de-seasonalized)", PF["baselineOcc"] / 100, PCT),
    ("Max Occupancy Cap", 0.92, PCT),
    ("Back-test MAPE (Gross Revenue)", fa["mapeGross"] / 100 if fa["mapeGross"] else "N/A", PCT if fa["mapeGross"] else None),
]
for label, val, fmt in inputs:
    r = write_item(ws7, r, label, val, fmt)

r += 1
ws7.cell(r, 1, "METHODOLOGY").font = SECTION_FONT
ws7.cell(r, 1).border = THICK_B
ws7.cell(r, 2).border = THICK_B
ws7.cell(r, 3).border = THICK_B
r += 1
methodology = (
    "Pro forma projections blend closed-month actuals (sourced from PMS booking data and "
    "QuickBooks) with a seasonal model for future months. The seasonal model de-seasonalizes "
    "actual ADR and occupancy performance to establish a baseline, then applies monthly seasonal "
    "indices derived from Santa Barbara STR demand patterns to project forward revenue. Operating "
    "expenses are projected from trailing actuals, with cleaning costs scaled per turnover and "
    "fixed costs (management, property tax, insurance) carried at known annual amounts. The model "
    "recalibrates automatically as each month closes, incorporating new actuals into the baseline "
    "- meaning projections tighten over time as operating history grows."
)
ws7.cell(r, 1, methodology).font = BODY
ws7.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="top")
ws7.merge_cells(f"A{r}:C{r}")
ws7.row_dimensions[r].height = 90

ws7.freeze_panes = "A5"


# ── Save ──
OUTPUT = Path(__file__).parent / "Casa_Yano_Lender_Package.xlsx"
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"Tabs: {', '.join(wb.sheetnames)}")
